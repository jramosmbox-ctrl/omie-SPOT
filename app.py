import io
import statistics
import datetime as dt
from collections import defaultdict

import streamlit as st

# Pagina --------------------------------------------------------------------
st.set_page_config(
    page_title="OMIE Portugal - Precos Horarios",
    page_icon="energy",
    layout="centered",
)

st.title(" OMIE Portugal - Precos Horarios + Analise BESS")
st.caption("Dados reais do mercado diario MIBEL  Fonte: www.omie.es")
st.divider()

# Formulario de datas -------------------------------------------------------
st.subheader(" Periodo de dados")

col1, col2 = st.columns(2)
with col1:
    data_ini = st.date_input(
        "Data de inicio",
        value=dt.date(2025, 1, 1),
        min_value=dt.date(2015, 1, 1),
        max_value=dt.date.today(),
        format="DD/MM/YYYY",
    )
with col2:
    data_fim = st.date_input(
        "Data de fim",
        value=dt.date(2025, 12, 31),
        min_value=dt.date(2015, 1, 1),
        max_value=dt.date.today(),
        format="DD/MM/YYYY",
    )

if data_fim < data_ini:
    st.error("  A data de fim tem de ser posterior a data de inicio.")
    st.stop()

n_dias = (data_fim - data_ini).days + 1
tempo_est = max(1, round(n_dias * 0.35 / 60, 1))
st.info(f"**{n_dias} dias** selecionados  Download estimado: ~{tempo_est} min")

st.divider()

# Parametros BESS -----------------------------------------------------------
st.subheader("Parametros BESS")

col3, col4, col5 = st.columns(3)
with col3:
    capacidade = st.number_input("Capacidade (MWh)", value=1.0, min_value=0.1, step=0.1)
    potencia   = st.number_input("Potencia (MW)",    value=0.5, min_value=0.1, step=0.1)
with col4:
    eficiencia = st.number_input("Eficiencia RT (%)", value=88, min_value=50, max_value=100) / 100
    opex       = st.number_input("OPEX (EUR/MWh/ano)",  value=8,  min_value=0,  step=1)
with col5:
    capex      = st.number_input("CAPEX (EUR/kWh)",     value=250, min_value=0, step=10)
    vida_util  = st.number_input("Vida util (anos)",  value=15,  min_value=1, max_value=30)
    wacc       = st.number_input("WACC (%)",           value=7,   min_value=1, max_value=30) / 100

st.divider()

# Botao principal -----------------------------------------------------------
if st.button("  Descarregar dados e gerar Excel", type="primary", use_container_width=True):

    # Importar aqui para nao bloquear o carregamento da pagina
    try:
        from OMIEData.DataImport.omie_marginalprice_importer import OMIEMarginalPriceFileImporter
    except ImportError:
        st.error("  Biblioteca OMIEData nao instalada. Verifica o requirements.txt.")
        st.stop()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        st.error("  Biblioteca openpyxl nao instalada. Verifica o requirements.txt.")
        st.stop()

    # Download --------------------------------------------------------------
    with st.spinner(f"A descarregar {n_dias} dias do OMIE (pode demorar {tempo_est} min)"):
        try:
            df_raw = OMIEMarginalPriceFileImporter(
                date_ini=dt.datetime.combine(data_ini, dt.time()),
                date_end=dt.datetime.combine(data_fim, dt.time()),
            ).read_to_dataframe(verbose=False)
        except Exception as e:
            st.error(f"  Erro ao descarregar dados: {e}")
            st.stop()

    df_pt = df_raw[df_raw["CONCEPT"] == "PRICE_PT"].copy()
    df_pt = df_pt.sort_values(by="DATE").reset_index(drop=True)

    if df_pt.empty:
        st.error("  Sem dados para o periodo selecionado.")
        st.stop()

    # Processar registos ----------------------------------------------------
    hora_cols = [f"H{i}" for i in range(1, 25)]
    records = []

    for _, row in df_pt.iterrows():
        d = row["DATE"]
        if hasattr(d, "date"):
            d = d.date()
        weekday = d.weekday()
        for h_idx, col in enumerate(hora_cols):
            if col not in row.index:
                continue
            try:
                price = float(row[col])
            except (ValueError, TypeError):
                continue
            h0 = h_idx
            hour_num = h_idx + 1
            records.append({
                "Data"      : str(d),
                "Hora"      : hour_num,
                "Hora_Label": f"{h0:02d}:00-{hour_num:02d}:00",
                "Timestamp" : f"{str(d)} {h0:02d}:00",
                "Preco"     : round(price, 2),
                "Mes"       : str(d)[:7],
                "DiaSemana" : ["Seg","Ter","Qua","Qui","Sex","Sab","Dom"][weekday % 7],
                "Periodo"   : (
                    "Vazio" if h0 in range(0, 7) else
                    "Cheia" if h0 in range(7, 10) or h0 in range(20, 24) else
                    "Ponta"
                ),
            })

    st.success(f"  {len(records)} registos horarios descarregados ({len(df_pt)} dias)")

    # Metricas rapidas ------------------------------------------------------
    todos_precos = [r["Preco"] for r in records]
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Media (EUR/MWh)",   f"{statistics.mean(todos_precos):.2f}")
    m2.metric("Min (EUR/MWh)",     f"{min(todos_precos):.2f}")
    m3.metric("Max (EUR/MWh)",     f"{max(todos_precos):.2f}")
    m4.metric("Horas negativas",  sum(1 for p in todos_precos if p < 0))

    # Helpers Excel ---------------------------------------------------------
    BLUE_DARK  = "1F3864"
    BLUE_MID   = "2E75B6"
    WHITE      = "FFFFFF"
    GREY_LIGHT = "F2F2F2"

    def hfill(h):
        return PatternFill("solid", start_color=h, fgColor=h)

    def bdr():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr_cell(ws, row, col, val, bg=BLUE_MID, bold=True, size=10,
                 color=WHITE, wrap=False, align="center"):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Arial", bold=bold, size=size, color=color)
        c.fill      = hfill(bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border    = bdr()
        return c

    def title_row(ws, merge, text, bg=BLUE_DARK, size=13):
        ws.merge_cells(merge)
        c = ws[merge.split(":")[0]]
        c.value     = text
        c.font      = Font(name="Arial", bold=True, size=size, color=WHITE)
        c.fill      = hfill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[int("".join(filter(str.isdigit, merge.split(":")[0])))].height = 28

    label = f"{data_ini.strftime('%d-%m-%Y')} a {data_fim.strftime('%d-%m-%Y')}"

    # Gerar Excel em memoria ------------------------------------------------
    with st.spinner("A gerar ficheiro Excel"):
        wb = Workbook()

        # Sheet 1 - Historico Horario
        ws1 = wb.active
        ws1.title = "Historico Horario"
        title_row(ws1, "A1:I1", f"OMIE - Precos Horarios Portugal  |  {label}  (EUR/MWh)", size=13)
        ws1.merge_cells("A2:I2")
        ws1["A2"].value = f"Fonte: OMIE - www.omie.es  |  {len(records)} registos horarios reais"
        ws1["A2"].font = Font(name="Arial", italic=True, size=9, color="595959")
        ws1["A2"].alignment = Alignment(horizontal="center")

        cols = [("Data",12),("Hora",7),("Periodo Horario",16),("Timestamp",20),
                ("Preco (EUR/MWh)",16),("Mes",10),("Dia Semana",12),
                ("Periodo Tarifario",18),("Ano",7)]
        for i,(h_,w) in enumerate(cols,1):
            hdr_cell(ws1,3,i,h_,wrap=True)
            ws1.column_dimensions[get_column_letter(i)].width = w
        ws1.row_dimensions[3].height = 30
        ws1.freeze_panes = "A4"

        for r_i, rec in enumerate(records, 4):
            fill = hfill(GREY_LIGHT) if r_i%2==0 else hfill(WHITE)
            row_data = [rec["Data"], rec["Hora"], rec["Hora_Label"], rec["Timestamp"],
                        rec["Preco"], rec["Mes"], rec["DiaSemana"], rec["Periodo"],
                        int(str(rec["Data"])[:4])]
            for c_i, val in enumerate(row_data, 1):
                cell = ws1.cell(row=r_i, column=c_i, value=val)
                cell.font = Font(name="Arial", size=9)
                cell.fill = fill; cell.border = bdr()
                cell.alignment = Alignment(horizontal="center")
                if c_i == 5:
                    cell.number_format = "#,##0.00"
                    if isinstance(val, float):
                        if val < 0:
                            cell.font = Font(name="Arial", size=9, color="FF0000", bold=True)
                        elif val > 100:
                            cell.font = Font(name="Arial", size=9, color="7030A0", bold=True)
                if c_i in (1,6):
                    cell.alignment = Alignment(horizontal="left")

        last = 3 + len(records)
        ws1.conditional_formatting.add(f"E4:E{last}", ColorScaleRule(
            start_type="min",  start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",    end_color="F8696B",
        ))

        # Sheet 2 - Resumo Mensal
        ws2 = wb.create_sheet("Resumo Mensal")
        title_row(ws2,"A1:L1", f"OMIE Portugal - Resumo Mensal (EUR/MWh)  |  {label}")
        monthly = defaultdict(list)
        for r in records: monthly[r["Mes"]].append(r["Preco"])
        month_names = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
        month_pt = {f"{y}-{m:02d}": month_names[m-1]+f"/{y}"
                    for y in range(data_ini.year, data_fim.year+1) for m in range(1,13)}

        hdrs2=[("Mes",12),("N. Horas",10),("Min (EUR/MWh)",13),("Max (EUR/MWh)",13),
               ("Media (EUR/MWh)",14),("Mediana (EUR/MWh)",15),("Desvio Padrao",14),
               ("Horas Negativas",16),("% Horas Neg.",14),
               ("Horas > 100EUR",13),("% Horas > 100EUR",15),("Spread Max-Min",15)]
        for i,(h_,w) in enumerate(hdrs2,1):
            hdr_cell(ws2,2,i,h_,wrap=True)
            ws2.column_dimensions[get_column_letter(i)].width=w
        ws2.row_dimensions[2].height=36

        all_prices=[]
        for r_i,m in enumerate(sorted(monthly),3):
            prices=monthly[m]; all_prices.extend(prices)
            n=len(prices); neg=sum(1 for p in prices if p<0); hi=sum(1 for p in prices if p>100)
            row=[month_pt.get(m,m),n,round(min(prices),2),round(max(prices),2),
                 round(statistics.mean(prices),2),round(statistics.median(prices),2),
                 round(statistics.stdev(prices) if n>1 else 0,2),
                 neg,neg/n,hi,hi/n,round(max(prices)-min(prices),2)]
            fill=hfill(GREY_LIGHT) if r_i%2==0 else hfill(WHITE)
            for c_i,val in enumerate(row,1):
                cell=ws2.cell(row=r_i,column=c_i,value=val)
                cell.font=Font(name="Arial",size=10); cell.fill=fill
                cell.border=bdr(); cell.alignment=Alignment(horizontal="center")
                if c_i in (3,4,5,6,7,12): cell.number_format="#,##0.00"
                if c_i in (9,11):         cell.number_format="0.0%"

        tr=3+len(monthly); n_a=len(all_prices)
        neg_a=sum(1 for p in all_prices if p<0); hi_a=sum(1 for p in all_prices if p>100)
        totals=["TOTAL",n_a,round(min(all_prices),2),round(max(all_prices),2),
                round(statistics.mean(all_prices),2),round(statistics.median(all_prices),2),
                round(statistics.stdev(all_prices),2),neg_a,neg_a/n_a,hi_a,hi_a/n_a,
                round(max(all_prices)-min(all_prices),2)]
        for c_i,val in enumerate(totals,1):
            cell=ws2.cell(row=tr,column=c_i,value=val)
            cell.font=Font(name="Arial",bold=True,size=10,color=WHITE)
            cell.fill=hfill(BLUE_DARK); cell.border=bdr()
            cell.alignment=Alignment(horizontal="center")
            if c_i in (3,4,5,6,7,12): cell.number_format="#,##0.00"
            if c_i in (9,11):         cell.number_format="0.0%"
        ws2.freeze_panes="A3"

        # Sheet 3 - Perfil Horario
        ws3 = wb.create_sheet("Perfil Horario")
        title_row(ws3,"A1:G1",f"OMIE Portugal - Perfil Horario Medio (EUR/MWh)  |  {label}")
        hourly=defaultdict(list)
        for r in records: hourly[r["Hora"]].append(r["Preco"])
        hdrs3=[("Hora",8),("Periodo",10),("Media (EUR/MWh)",14),("Min (EUR/MWh)",13),
               ("Max (EUR/MWh)",13),("Mediana (EUR/MWh)",15),("Desvio Padrao",14)]
        for i,(h_,w) in enumerate(hdrs3,1):
            hdr_cell(ws3,2,i,h_,wrap=True)
            ws3.column_dimensions[get_column_letter(i)].width=w
        ws3.row_dimensions[2].height=30
        PC={"Vazio":"EBF3FB","Cheia":"FFF2CC","Ponta":"FCE4D6"}
        periodo=lambda h:("Vazio" if (h-1) in range(0,7) else
                          "Cheia" if (h-1) in range(7,10) or (h-1) in range(20,24) else "Ponta")
        for h in range(1,25):
            prices=hourly.get(h,[0]); p=periodo(h)
            row=[f"{(h-1):02d}:00",p,round(statistics.mean(prices),2),round(min(prices),2),
                 round(max(prices),2),round(statistics.median(prices),2),
                 round(statistics.stdev(prices) if len(prices)>1 else 0,2)]
            for c_i,val in enumerate(row,1):
                cell=ws3.cell(row=h+2,column=c_i,value=val)
                cell.font=Font(name="Arial",size=10); cell.fill=hfill(PC[p])
                cell.border=bdr(); cell.alignment=Alignment(horizontal="center")
                if c_i in (3,4,5,6,7): cell.number_format="#,##0.00"

        # Sheet 4 - Analise BESS
        ws4 = wb.create_sheet("Analise BESS")
        title_row(ws4,"A1:K1",f"Analise de Arbitragem BESS - Portugal  |  {label}",size=13)
        ws4.merge_cells("A3:D3"); hdr_cell(ws4,3,1,"PARAMETROS DO SISTEMA BESS",size=11)
        for c in range(2,5): ws4.cell(row=3,column=c).fill=hfill(BLUE_MID)
        params_vals=[("Capacidade (MWh)",capacidade,"MWh"),("Potencia (MW)",potencia,"MW"),
                     ("Eficiencia RT",eficiencia,"%"),("OPEX (EUR/MWh/ano)",opex,"EUR/MWh"),
                     ("CAPEX (EUR/kWh)",capex,"EUR/kWh"),("Vida util (anos)",vida_util,"anos"),
                     ("WACC",wacc,"%")]
        for i,lbl in enumerate(["Parametro","Valor","Unidade"],1):
            hdr_cell(ws4,4,i,lbl,bg="4472C4",size=9)
        for r_i,(name,val,unit) in enumerate(params_vals,5):
            fill=hfill(GREY_LIGHT) if r_i%2==0 else hfill(WHITE)
            for c_i,v in enumerate([name,val,unit],1):
                cell=ws4.cell(row=r_i,column=c_i,value=v)
                cell.font=Font(name="Arial",size=9,color="0000FF" if c_i==2 else "000000",bold=(c_i==2))
                cell.fill=fill; cell.border=bdr(); cell.alignment=Alignment(horizontal="center")
        ws4.column_dimensions["A"].width=30; ws4.column_dimensions["B"].width=12; ws4.column_dimensions["C"].width=14

        ws4.merge_cells("E3:K3"); hdr_cell(ws4,3,5,"ARBITRAGEM DIARIA - 1 CICLO/DIA",size=11)
        for c in range(6,12): ws4.cell(row=3,column=c).fill=hfill(BLUE_MID)
        arb_hdrs=[("Data",12),("Min (EUR/MWh)",13),("H. Carga",9),("Max (EUR/MWh)",13),
                  ("H. Descarga",11),("Spread (EUR/MWh)",14),("Receita Bruta (EUR)",16)]
        for i,(h_,w) in enumerate(arb_hdrs,5):
            hdr_cell(ws4,4,i,h_,bg="4472C4",size=9,wrap=True)
            ws4.column_dimensions[get_column_letter(i)].width=w
        ws4.row_dimensions[4].height=30

        day_recs=defaultdict(list)
        for rec in records: day_recs[rec["Data"]].append(rec)
        arb_rows=[]
        for d_str in sorted(day_recs):
            recs=day_recs[d_str]
            min_r=min(recs,key=lambda x:x["Preco"]); max_r=max(recs,key=lambda x:x["Preco"])
            spread=max_r["Preco"]-min_r["Preco"]; rev=spread*potencia*eficiencia
            arb_rows.append((d_str,min_r["Preco"],min_r["Hora"],max_r["Preco"],max_r["Hora"],spread,rev))

        for r_i,row in enumerate(arb_rows,5):
            fill=hfill(GREY_LIGHT) if r_i%2==0 else hfill(WHITE)
            for c_i,val in enumerate(row,5):
                cell=ws4.cell(row=r_i,column=c_i,value=val)
                cell.font=Font(name="Arial",size=9); cell.fill=fill
                cell.border=bdr(); cell.alignment=Alignment(horizontal="center")
                if c_i in (6,7,8,11): cell.number_format="#,##0.00"
                if c_i==10 and isinstance(val,float) and val>50:
                    cell.font=Font(name="Arial",size=9,color="00B050",bold=True)

        total_rev=sum(r[6] for r in arb_rows); avg_spread=statistics.mean(r[5] for r in arb_rows)
        total_opex=opex*capacidade; capex_tot=capex*capacidade*1000
        annuity=capex_tot*(wacc*(1+wacc)**vida_util)/((1+wacc)**vida_util-1)
        ebitda=total_rev-total_opex

        sr=5+len(params_vals)+3
        ws4.merge_cells(f"A{sr}:D{sr}"); hdr_cell(ws4,sr,1,"RESUMO FINANCEIRO",size=11)
        for c in range(2,5): ws4.cell(row=sr,column=c).fill=hfill(BLUE_MID)
        summary=[("Receita Bruta Total (EUR)",total_rev,"EUR"),("OPEX Anual (EUR)",total_opex,"EUR"),
                 ("EBITDA (EUR)",ebitda,"EUR"),("CAPEX Total (EUR)",capex_tot,"EUR"),
                 ("Anuidade CAPEX (EUR/ano)",annuity,"EUR/ano"),
                 ("Cash Flow Liquido Anual (EUR)",ebitda-annuity,"EUR"),
                 ("Spread Medio Diario (EUR/MWh)",avg_spread,"EUR/MWh"),
                 ("Dias com Spread > 50 EUR/MWh",sum(1 for r in arb_rows if r[5]>50),"dias"),
                 ("Dias com Spread > 100 EUR/MWh",sum(1 for r in arb_rows if r[5]>100),"dias")]
        for i,(lbl,val,unit) in enumerate(summary,sr+1):
            fill=hfill(GREY_LIGHT) if i%2==0 else hfill(WHITE)
            ws4.cell(row=i,column=1,value=lbl).font=Font(name="Arial",size=9)
            ws4.cell(row=i,column=1).fill=fill; ws4.cell(row=i,column=1).border=bdr()
            cell_v=ws4.cell(row=i,column=2,value=round(val,2))
            cell_v.font=Font(name="Arial",size=9,bold=True); cell_v.number_format="#,##0.00"
            cell_v.fill=fill; cell_v.border=bdr(); cell_v.alignment=Alignment(horizontal="center")
            ws4.cell(row=i,column=3,value=unit).font=Font(size=9)
            ws4.cell(row=i,column=3).fill=fill; ws4.cell(row=i,column=3).border=bdr()

        # Guardar em memoria
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

    # Botao de download -----------------------------------------------------
    nome_ficheiro = (
        f"OMIE_Portugal_"
        f"{data_ini.strftime('%d%m%Y')}_"
        f"{data_fim.strftime('%d%m%Y')}_"
        f"Precos_Horarios_BESS.xlsx"
    )

    st.download_button(
        label="  Descarregar Excel",
        data=buf,
        file_name=nome_ficheiro,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

    st.caption(f"Ficheiro: `{nome_ficheiro}`    4 folhas    {len(records)} registos horarios")
